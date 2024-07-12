import pyodbc  # Библиотека подключения Access
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import tkinter as tk  # Библиотека для интерфейса
from tkinter import filedialog
from tkinter import ttk
from tkinter import Tk, Label, Button, Toplevel, StringVar, OptionMenu, messagebox
from datetime import datetime, date
from tkcalendar import DateEntry
from sqlalchemy import create_engine

SVovlech_dict = {
    1: 'подлежат включению в границы населенного пункта для его развития',
    2: 'подлежат вовлечению в сельскохозяйственный оборот',
    3: 'подлежат вовлечению в лесохозяйственный оборот',
    4: 'подлежат вовлечению для использования в иных целях',
    5: 'включены в границы населенного пункта для его развития',
    6: 'вовлечены в сельскохозяйственный оборот',
    7: 'вовлечены в лесохозяйственный оборот',
    8: 'вовлечены для использования в иных целях',
    9: 'не могут быть использованы в хозяй-ственной деятельности',
    0: 'Не обследовано местным исполнительным комитетом'
}
Forma22_dict = {
    "01": "01 - сельскохозяйственные организации, использующие предоставленные им земли для ведения сельского хозяйства, в том числе в исследовательских и учебных целях, а также для ведения подсобного хозяйства",
    "02": "02 - сельскохозяйственные организации Министерства сельского хозяйства и продовольствия Республики Беларусь",
    "03": "03 - крестьянские (фермерские) хозяйства",
    "05": "05 - граждане, использующие земельные участки для строительства и (или) обслуживания жилого дома",
    "06": "06 - граждане, использующие земельные участки для ведения личного подсобного хозяйства",
    "07": "07 - граждане, использующие земельные участки для садоводства и дачного строительства",
    "08": "08 - граждане, использующие земельные участки для огородничества",
    "09": "09 - граждане, использующие земельные участки для сенокошения и выпаса сельскохозяйственных животных",
    "10": "10 - граждане, использующие земельные участки для иных сельскохозяйственных целей",
    "11": "11 - граждане, использующие земельные участки для иных несельскохозяйственных целей",
    "12": "12 - промышленные организации",
    "13": "13 - организации железнодорожного транспорта",
    "14": "14 - -организации автомобильного транспорта",
    "15": "15 - организации Вооруженных Сил Республики Беларусь, воинских частей, военных учебных заведений и других войск и воинских формирований Республики Беларусь",
    "16": "16 - организации воинских частей, военных учебных заведений и других войск и воинских формирований иностранных государств",
    "17": "17 - организации связи, энергетики, строительства, торговли, образования, здравоохранения и иные землепользователи",
    "18": "18 - организации природоохранного, оздоровительного, рекреационного и историко - культурного назначения",
    "19": "19 - заповедники, национальные парки и дендрологические парки",
    "20": "20 - организации, ведущие лесное хозяйство",
    "21": "21 - организации, эксплуатирующие и обслуживающие гидротехнические и иные водохозяйственные сооружения",
    "22": "22 - -земли, земельные участки, не предоставленные землепользователям",
    "23": "23 - земли общего пользования в населенных пунктах, садоводческих товариществах и дачных кооперативах, а также земельные участки, используемые гражданами",
    "24": "24 - иные земли общего пользования за пределами границ населенных пунктов"
}




#"Брестская обл.", "Витебская обл.", "Гомельская обл.", "Гродненская обл.", "Минская обл.", "Могилевская обл."
def clear_sheet(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None


def update_spreadsheet(connection_string, path: str, _df, startcol: int = 1, startrow: int = 1,
                       sheet_name: str = "ToUpdate"):
    '''
    :param path: Путь до файла Excel
    :param _df: Датафрейм Pandas для записи
    :param startcol: Стартовая колонка в таблице листа Excel, куда буду писать данные
    :param startrow: Стартовая строка в таблице листа Excel, куда буду писать данные
    :param sheet_name: Имя листа в таблице Excel, куда буду писать данные
    :return:
    '''
    engine = create_engine(f"access+pyodbc:///?odbc_connect={connection_string}")

    wb = openpyxl.load_workbook(path)
    sheet = wb[sheet_name]
    clear_sheet(sheet)  # Предполагается, что функция clear_sheet определена где-то в вашем коде

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    ic = 0
    ir = 0
    for ir_Rayon in range(0, len(_df)):
        sheet.cell(startrow + ir, startcol + ic).value = _df.iloc[ir_Rayon, 0]  # область
        sheet.cell(startrow + ir + 1, startcol + ic).value = _df.iloc[ir_Rayon, 1]  # район
        RayonForSearch = _df.iloc[ir_Rayon, ic + 1]  # Переменная района для запроса
        print(RayonForSearch)

        queryRayon = '''
                     SELECT Forma22, COUNT(OBJECTID) as [Кол-во контуров], SUM(Shape_Area) as [Площадь га]
        FROM dkr_table
        where Rayon = '%s'
        GROUP BY Forma22
                    ''' % (RayonForSearch)  # Общие формы для района
        querySVovlech = '''
        SELECT SVovlech, COUNT(OBJECTID) as [Кол-во контуров], SUM(Shape_Area) as [Площадь га]
        FROM dkr_table
        where Rayon = '%s'
        GROUP BY SVovlech
                        ''' % (RayonForSearch)  # Общая вовлеченность для района
        _dfRayon = pd.read_sql_query(queryRayon, engine)
        _dfSVovlech = pd.read_sql_query(querySVovlech, engine)

        sheet.cell(startrow + ir + 1, startcol + ic + 2).value = _df.iloc[ir_Rayon, 2]  # общие контура
        sheet.cell(startrow + ir + 1, startcol + ic + 3).value = _df.iloc[ir_Rayon, 3]  # общая площадь

        for i in range(0, len(_dfRayon)):
            sheet.cell(startrow + ir + i + 2, 2).value = Forma22_dict.get(_dfRayon.iloc[i, 0])  # Номер формы
            sheet.cell(startrow + ir + i + 2, 3).value = _dfRayon.iloc[i, 1]  # контура формы
            sheet.cell(startrow + ir + i + 2, 4).value = _dfRayon.iloc[i, 2]  # площадь формы
            Forma22Current = _dfRayon.iloc[i, 0]

            queryForma22 = '''
             SELECT  SVovlech, COUNT(OBJECTID) as [Кол-во контуров], SUM(Shape_Area) as [Площадь га]
            FROM dkr_table
            where Rayon = '%s' AND Forma22 = '%s'
            GROUP BY SVovlech''' % (RayonForSearch, Forma22Current)  # Перекрестие Свовлеченности и формы

            _dfForma22 = pd.read_sql_query(queryForma22, engine)
            for j in range(0, len(_dfForma22)):
                sheet.cell(startrow + ir + i + 2, _dfForma22.iloc[j, 0] * 2 + 3).value = _dfForma22.iloc[
                    j, 1]  # контура формы
                sheet.cell(startrow + ir + i + 2, _dfForma22.iloc[j, 0] * 2 + 4).value = _dfForma22.iloc[
                    j, 2]  # площадь формы

        for i in range(0, len(_dfSVovlech)):
            sheet.cell(startrow + ir + 1, _dfSVovlech.iloc[i, 0] * 2 + 3).value = _dfSVovlech.iloc[
                i, 1]  # контура формы
            sheet.cell(startrow + ir + 1, _dfSVovlech.iloc[i, 0] * 2 + 4).value = _dfSVovlech.iloc[
                i, 2]  # площадь формы

        ir += len(_dfRayon) + 2

    # Добавление границ к ячейкам
    max_row = sheet.max_row
    max_col = sheet.max_column
    for row in sheet.iter_rows(min_row=startrow, max_row=max_row, min_col=startcol, max_col=max_col):
        for cell in row:
            cell.border = thin_border

    wb.save(path)

# Функция выбора файла
def select_database_file():
    root = tk.Tk()
    root.withdraw()

    # Диалоговое окно выбора файла
    file_path = filedialog.askopenfilename(
        title="Выберите файл Access",
        filetypes=(("Access files", "*.accdb;*.mdb"), ("All Files", "*.*"))
    )
    return file_path

# Функция формирования статичных полей таблицы
def mask_table():
    alignment = openpyxl.styles.Alignment(wrap_text=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    workbook = openpyxl.load_workbook('FinalTable.xlsx')
    sheet = workbook.worksheets[0]

    # Установка значений и объединение ячеек
    sheet.merge_cells('A1:Z1')
    sheet[
        "A1"].value = "ИНФОРМАЦИЯ О ВОВЛЕЧЕНИИ ЗЕМЕЛЬ ПОД ДРЕСНО-КУССТАРНИКОВОЙ РАСТИТЕЛЬНОСТЬЮ В ХОЗЯЙСТВЕННЫЙ ОБОРОТ"

    sheet["A3"].value = "За период"
    sheet["A4"].value = "Область"

    sheet.merge_cells('A7:A10')
    sheet["A7"].value = "Наименование"

    sheet.merge_cells('E7:Z7')
    sheet["E7"].value = "Из них"

    sheet.merge_cells('E8:V8')
    sheet["E8"].value = "По результатам"

    sheet.merge_cells('C7:D9')
    sheet["C7"].value = "Всего земель"

    sheet.merge_cells('W8:X9')
    sheet["W8"].value = "Обследовано"

    sheet.merge_cells('Y8:Z9')
    sheet["Y8"].value = "Не обследовано"

    sheet.merge_cells('B7:B10')
    sheet["B7"].value = "Категория"

    # Заполнение и выравнивание ячеек
    for i in range(3, 27, 2):
        sheet.cell(row=10, column=i).value = "количество контуров"
        sheet.cell(row=10, column=i).alignment = alignment
        sheet.cell(row=10, column=i + 1).value = "площадь га"
        sheet.cell(row=10, column=i + 1).alignment = alignment

    for i in range(1, 27):
        sheet.cell(row=11, column=i).value = i

    j = 1
    for i in range(5, 23, 2):
        sheet.merge_cells(start_row=9, start_column=i, end_row=9, end_column=i + 1)
        sheet.cell(row=9, column=i).value = SVovlech_dict.get(j)
        sheet.cell(row=9, column=i).alignment = alignment
        j += 1

    # Добавление границ к ячейкам
    for row in sheet.iter_rows(min_row=1, max_row=11, min_col=1, max_col=26):
        for cell in row:
            cell.border = thin_border

    workbook.save('FinalTable.xlsx')


def connect_to_database(file_path):
    if not file_path:
        print("Файл не выбран")
        return

    def on_submit():
        start_date = start_date_entry.get_date()
        end_date = end_date_entry.get_date()
        region = region_combobox.get()

        if start_date > end_date:
            messagebox.showerror("Ошибка", "Начальная дата не может быть позже конечной даты.")
            return

        selected_values['start_date'] = start_date
        selected_values['end_date'] = end_date
        selected_values['region'] = region
        root.quit()

    selected_values = {'start_date': None, 'end_date': None, 'region': None}

    root = tk.Tk()
    root.title("Выбор даты и области")

    ttk.Label(root, text="Начальная дата:").grid(column=0, row=0, padx=10, pady=5)
    start_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2,
                                 date_pattern='dd/mm/yyyy')
    start_date_entry.grid(column=1, row=0, padx=10, pady=5)

    ttk.Label(root, text="Конечная дата:").grid(column=0, row=1, padx=10, pady=5)
    end_date_entry = DateEntry(root, width=12, background='darkblue', foreground='white', borderwidth=2,
                               date_pattern='dd/mm/yyyy')
    end_date_entry.grid(column=1, row=1, padx=10, pady=5)


    ttk.Label(root, text="Область:").grid(column=0, row=2, padx=10, pady=5)
    regions = ["Брестская обл.", "Витебская обл.", "Гомельская обл.", "Гродненская обл.", "Минская обл.", "Могилевская обл."]
    region_combobox = ttk.Combobox(root, values=regions, state="readonly")
    region_combobox.grid(column=1, row=2, padx=10, pady=5)
    region_combobox.current(0)

    submit_button = ttk.Button(root, text="Подтвердить", command=on_submit)
    submit_button.grid(column=0, row=3, columnspan=2, pady=10)

    root.mainloop()
    root.destroy()

    # Вызов функции для отображения диалогового окна и получения значений
    start_date = selected_values['start_date']
    end_date = selected_values['end_date']
    region = selected_values['region']
    print(f"Начальная дата:", start_date)
    print(f"Конечная дата:", end_date)
    print(f"Выбранная область:", region)
    con_str = rf"DBQ={file_path};"
    print(pyodbc.drivers())

    print(con_str)
    # Формирование строки подключения
    connection_string = (
            r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};'
            rf"DBQ={file_path};"
    )

    try:
        engine = create_engine(f"access+pyodbc:///?odbc_connect={connection_string}")
        print(region, 'gbgbg')
        query = '''
         SELECT Oblast as [Область], Rayon as [Район], COUNT(OBJECTID) as [Количество контуров],
         SUM(Shape_Area) as [Площадь га]
FROM dkr_table
where Oblast = '%s'
GROUP BY Oblast, Rayon;
        ''' % (region)
        df = pd.read_sql_query(query,engine)
        df.to_excel("FinalTable.xlsx", index=False, startrow=5, startcol=1,header=False)  # Костыль очистки таблицы
        update_spreadsheet( connection_string, "FinalTable.xlsx", df, 1, 12, "Sheet1")
    except pyodbc.Error as e:
        print("Ошибка соединения: ", e)

    finally:
        engine.dispose()

if __name__ == "__main__":
    # Проверка драйверов
    drivers = pyodbc.drivers()
    print("Включенные ODBC драйвера:")
    for driver in drivers:
        print(driver)
    file_path = select_database_file()
    connect_to_database(file_path)
    mask_table()

